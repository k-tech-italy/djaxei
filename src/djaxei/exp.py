import os
from collections import OrderedDict
from tempfile import NamedTemporaryFile

from django.contrib.admin.utils import NestedObjects
from django.db import router
from django.db.models.query import QuerySet
from openpyxl import Workbook

from djaxei.providers import get_workbook_impl


class Exporter:
    def __init__(self, root, rules: dict):
        """Create an exporter.

        :param root: Can be a Django object, or a queryset
        :param rules: A dictionary 'appname.modelname': list of field names or tuple(fieldname, stringify function)
        """
        if isinstance(root, QuerySet):
            self.roots = root
            self.using = router.db_for_write(root.first()._meta.model)
        else:
            self.roots = [root]
            self.using = router.db_for_write(root._meta.model)
        self.rules = rules

    def xls_export(self, target):
        lmodels = {}

        wb = Workbook(write_only=True)
        sheets = OrderedDict()

        for model_ref, field_refs in self.rules.items():
            model_ref = model_ref.lower()

            lmodels[model_ref] = field_refs

            header = [x if isinstance(x, str) else x[0] for x in field_refs]
            sheets[model_ref] = [header, ]
            wb.create_sheet(model_ref)

        collector = NestedObjects(using=self.using)
        collector.collect(self.roots)

        def callback(obj):
            fields = lmodels.get(obj._meta.label_lower, None)
            if fields:
                row = []
                for x in fields:
                    if isinstance(x, str):
                        row.append(getattr(obj, x))
                    else:
                        row.append(x[1](getattr(obj, x[0])))
                sheets[obj._meta.label_lower].append(row)

        collector.nested(callback)

        for k, rows in sheets.items():
            for row in rows:
                wb[k].append(row)

        wb.save(target)
        #
        # workbook = None
        # workbookfile = None
        # try:
        #     sheets = OrderedDict()
        #
        #     lmodels = {}
        #     for k, v in _models.items():
        #         lname = k.lower()
        #         model_name = lname.rsplit('.')[1]
        #         lmodels[lname] = v
        #         sheets[model_name] = [v, ]
        #
        #     if root:
        #         root_qs = root._meta.model.objects.filter(pk=root.pk)
        #
        #     using = router.db_for_write(root_qs.first()._meta.model)
        #     collector = NestedObjects(using=using)
        #     collector.collect(root_qs)
        #
        #     def callback(obj):
        #         fields = lmodels.get(obj._meta.label_lower, None)
        #         if fields:
        #             sheets[obj._meta.model_name].append([getattr(obj, x) for x in fields])
        #
        #     collector.nested(callback)
        #
        #     Workbook = get_workbook_impl()
        #     workbookfile = self.dest or NamedTemporaryFile(dir=self.tmpdir, suffix=Workbook._preferred_suffix, delete=False)
        #     Workbook(workbookfile).write_data(sheets)
        #
        #     return workbookfile.name
        #
        # except Exception as e:
        #     if workbook:
        #         if not workbookfile.closed:
        #             workbookfile.close()
        #         if os.path.exists(workbookfile.name):
        #             os.remove(workbookfile.name)
        #     raise e